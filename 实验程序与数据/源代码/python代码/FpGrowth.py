import os
import time
import openpyxl

class FPNode:
    def __init__(self, item, count, parent):
        self.item = item        # 项的名称
        self.count = count      # 项出现的次数
        self.parent = parent    # 父亲指针,指向上一个节点
        self.next = None        # 项头表中相同项的指针
        self.children = {}      # 儿子指针,指向所有儿子节点

    def display(self, ind=1):   # 输出当前节点的基本信息
        print(''*ind, self.item, '', self.count)
        for child in self.children.values():
            child.display(ind+1)

class FPgrowth:
    # 初始化最小支持度和最小置信度,此处minsup是出现次数,而非概率
    # 改support,min_confidence,k
    def __init__(self, support, min_confidence,k):
        # self.support = support
        self.min_support = support
        self.min_confidence = min_confidence
        self.Maxk = k-1
        self.cnt = 0

    def transfer2FrozenDataSet(self, data):
        frozen_data = {}
        for elem in data:
            frozen_data[frozenset(elem)] = 1
        return frozen_data


    def updataTree(self, data, FP_tree, header, count):
        frequent_item = data[0] # 取出当前项集的第一项
        if frequent_item in FP_tree.children: # 如果该项已经在树中
            FP_tree.children[frequent_item].count += count # 累加上出现次数
        else:# 如果不再树中,则需要新建一个儿子节点插入到当前节点之下
            FP_tree.children[frequent_item] = FPNode(frequent_item, count, FP_tree) # 新建节点
            # 判断该项是否与项头表连接
            if header[frequent_item][1] is None:  # 没连直接连接即可
                header[frequent_item][1] = FP_tree.children[frequent_item]
            else: # 连接过了的话要找到 指针末尾再连接
                self.updateHeader(header[frequent_item][1], FP_tree.children[frequent_item]) # share the same path
        
        # 继续递归处理当前事务的后面项
        if len(data) > 1:
            self.updataTree(data[1::], FP_tree.children[frequent_item], header, count)  # recurrently update FP tree


    def updateHeader(self, head_node, tail_node):
        while head_node.next is not None:
            head_node = head_node.next  # 项头表指针一直往前走
        head_node.next = tail_node      # 走到末尾将要加入的指针连接上去

    def createFPTree(self, train_data):
        initial_header = {}

        # 第一遍扫描,获取单个项的计数
        for record in train_data:       # 枚举每一个事务
            for item in record:         # 枚举每一项
            # 获得项出现的次数
                initial_header[item] = initial_header.get(item, 0) + train_data[record]

        # 剔除低于最小支持度的项,并用header储存剩余项
        header = {}
        for k in initial_header.keys():
            if initial_header[k]/self.cnt >= self.min_support:
                header[k] = initial_header[k]
        # 转为set类型   
        frequent_set = set(header.keys())
        # 如果不存在频繁项集直接 返回None
        if len(frequent_set) == 0:
            return None, None

        # 对header进行扩充,原有header[k]仅表示k出现的次数
        # 现增加新的一部分用于储存指针,此后header[k][1]表示与k对应的项头表指针
        for k in header:
            header[k] = [header[k], None]

        # 第二遍扫描,建立FP Tree
        FP_tree = FPNode('root', 1, None)      # 创建根节点
        for record, count in train_data.items():# 枚举每一个事务
            frequent_item = {}                 # 先将低于最小支持度的剔除
            for item in record:                # 枚举每一项
                if item in frequent_set:       # 如果在频繁项集里(即大于最小支持度),先存进去
                    frequent_item[item] = header[item][0]
            
            if len(frequent_item) > 0:
                # 按出现次数从大到小排序
                ordered_frequent_item = [val[0] for val in sorted(frequent_item.items(), key=lambda val:val[1], reverse=True)]
                # 将项按次序插入FP Tree
                self.updataTree(ordered_frequent_item, FP_tree, header, count)

        return FP_tree, header

    def ascendTree(self, node):
        prefix_path = []
        while node.parent != None and node.parent.item != 'root':
        # 如果父亲节点存在且不为根节点
            node = node.parent              # 往上走
            prefix_path.append(node.item)   # 并把经过节点的名字加入路径
        return prefix_path

    def getPrefixPath(self, base, header):
        prefix_path = {}
        start_node = header[base][1]

        prefixs = self.ascendTree(start_node)
        if len(prefixs) != 0:
            # 设为起始节点的 count
            prefix_path[frozenset(prefixs)] = start_node.count
        # 如果存在多条路径,即项头表中base节点存在指针
        while start_node.next is not None:
            start_node = start_node.next            # 跟着指针走
            prefixs = self.ascendTree(start_node)   # 获得路径
            if len(prefixs) != 0:
                prefix_path[frozenset(prefixs)] = start_node.count  # 累加出现次数
        return prefix_path

    def findFrequentItem(self, header, prefix, frequent_set):
        # 对项头表按出现次数从大到小排序
        if len(prefix) <= self.Maxk:
            header_items = [val[0] for val in sorted(header.items(), key=lambda val: val[1][0])]
            if len(header_items) == 0:
                return
            for base in header_items:
                new_prefix = prefix.copy()
                new_prefix.add(base)
                # 记录当前路径和出现次数
                support = header[base][0]
                frequent_set[frozenset(new_prefix)] = support
                # 获得前缀路径
                prefix_path = self.getPrefixPath(base, header)
                if len(prefix_path) != 0:
                    # 创建新的FP Tree
                    conditonal_tree, conditional_header = self.createFPTree(prefix_path)
                    if conditional_header is not None:
                        # 递归发掘频繁项
                        self.findFrequentItem(conditional_header, new_prefix, frequent_set)

    def generateRules(self, frequent_set, rules):
        for frequent_item in frequent_set:
            if len(frequent_item) > 1:
                self.getRules(frequent_item, frequent_item, frequent_set, rules)

    def removeItem(self, current_item, item):
        tempSet = []
        for elem in current_item:
            if elem != item:
                tempSet.append(elem)
        tempFrozenSet = frozenset(tempSet)
        return tempFrozenSet

    def getRules(self, frequent_item, current_item, frequent_set, rules):
        for item in current_item:
            subset = self.removeItem(current_item, item)
            # 去除item以计算置信度
            # 置信度 = 共同发生的概率 / 去除某个事件后发生的概率 例如 P(AB)/P(A)
            if subset in frequent_set:
                confidence = frequent_set[frequent_item]/frequent_set[subset]
            else:
                confidence = 0
            # 如果大于最小置信度就存起来
            if confidence >= self.min_confidence:
                flag = False
                for rule in rules: # 判断是否已经存在相同规则
                    if (rule[0] == subset) and (rule[1] == frequent_item - subset):
                        flag = True

                if flag == False: # 没有则加入
                    rules.append((subset, frequent_item - subset, confidence))

                if (len(subset) >= 2):# 递归挖掘更多规则
                    self.getRules(frequent_item, subset, frequent_set, rules)

    # 相当于主函数,FpTree算法完整过程
    def train(self, data, display=True):

        data = self.transfer2FrozenDataSet(data)        # 载入数据
        FP_tree, header = self.createFPTree(data)       # 建树
        # print(self.cnt)
        # FP_tree.display()
        frequent_set = {}
        prefix_path = set([])
        # 挖掘频繁项集
        self.findFrequentItem(header, prefix_path, frequent_set)
        rules = []
        # 寻找关联规则
        self.generateRules(frequent_set, rules)
        # 输出结果
        if display:
            workbook = openpyxl.load_workbook('fpgrowth.xlsx')
            sheet = workbook['Sheet1']
            for i in range(len(rules)):
                tmpcnt = 1
                for x in range(0,2):
                    for j in rules[i][x]:
                        sheet.cell(i+1,tmpcnt).value = str(j)
                        tmpcnt = tmpcnt+1
                sheet.cell(i+1,tmpcnt).value = rules[i][2]
            workbook.save('fpgrowth.xlsx')
            
        return frequent_set, rules

if __name__ == "__main__":
    # 设置最小支持度,置信度,挖掘项数
    tmp = FPgrowth(0.01,0.9,3)
    # path引号内的改为处理后数据集的路径
    path = r"C:\Users\zbwer\Desktop\md\关联规则\source\超市关联规则数据集"
    os.chdir(path)
    # 加载数据集
    workbook = openpyxl.load_workbook('dataset.xlsx')
    sheet = workbook['Sheet3']
    rows= sheet.max_row+1
    cols = sheet.max_column+1
    NewDa = []
    for i in range(1,rows):
        tmpD = []
        for j in range(1,cols):
            celldate = sheet.cell(i,j).value
            if celldate:
                tmpD.append(celldate)
        NewDa.append(tmpD)
    tmp.cnt = float(len(NewDa))
    # 挖掘关联规则
    E1 = time.time()
    tmp.train(NewDa)
    # 记录程序运行时间
    E2 = time.time()
    print(E2-E1)