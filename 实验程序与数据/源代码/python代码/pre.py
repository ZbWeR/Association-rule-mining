# 本程序用于处理原始数据
import os
import openpyxl
import xlwt

# path设置为原始数据所在的文件夹
path = r"C:\Users\zbwer\Desktop\md\关联规则\source\超市关联规则数据集"
os.chdir(path) # 改变当前工作目录到指定路径

workbook = openpyxl.load_workbook('Original data.xlsx')
sheet = workbook['Sheet1']#打开存储数据对应的sheet
rows= sheet.max_row+1
cols = sheet.max_column+1

# 读入数据
NewDa = []
for i in range(3,rows):#枚举行
    tmpDa = []
    for j in range(2,cols):#枚举列
        ch = sheet.cell(i,j).value
        if ch == 'T':#如果为T就加入商品
            name = sheet.cell(2,j).value
            tmpDa.append(name)
    NewDa.append(tmpDa)
# 处理后的数据保存在sheet3
sheet = workbook['Sheet3']
lincnt = 1
for i in range(0,len(NewDa)-1):
    cnt = 1
    if  len(NewDa[i]) >= 1:
        for x in NewDa[i]:
            sheet.cell(lincnt,cnt).value = x
            cnt = cnt+1
        lincnt =lincnt+1
workbook.save('dataset.xlsx')
# 处理后的数据保存在dataset.xlsx