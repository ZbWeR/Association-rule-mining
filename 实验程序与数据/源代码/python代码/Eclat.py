import os
import openpyxl
import time

class Eclat:
    def __init__(self,min_support, min_confidence,k):#改min_support,min_confidence
        self.min_support = min_support
        self.min_confidence = min_confidence
        self.cnt = 0
        self.Num = k

    def invert(self, data):
        invert_data = {}
        frequent_item = []
        support = []
        # 倒排,记录每个项出现的事务TID,用invert_data存储
        # invert_data[i] 表示 包含 i 这个项 的 所有事务
        for i in range(len(data)):
            for item in data[i]:
                if invert_data.get(item) is not None:
                    invert_data[item].append(i)
                else:
                    invert_data[item] = [i]
        # 删除小于最小支持度的
        for item in invert_data.keys():
            if len(invert_data[item])/self.cnt >= self.min_support:
                frequent_item.append([item])        # 加入频繁集
                support.append(invert_data[item])   # 存放支持度
        frequent_item = list(map(frozenset, frequent_item))
        return frequent_item, support

    def getIntersection(self, frequent_item, support):
        sub_frequent_item = []
        sub_support = []
        k = len(frequent_item[0]) + 1
        if k <= self.Num :
        # 枚举频繁集中的项 进行合并操作
            for i in range(len(frequent_item)): 
                for j in range(i+1, len(frequent_item)):
                    L1 = list(frequent_item[i])[:k-2]
                    L2 = list(frequent_item[j])[:k-2]
                    # 如果前k-1项相同,才能 取并集 获得k项频繁集
                    # 这里是k-2, 是因为对于含有n个元素的list,其下表范围为[0,n-1]
                    if L1 == L2:
                        flag = len(list(set(support[i]).intersection(set(support[j]))))
                        # flag是新添加的项 的出现次数,如果大于最小支持度就加进去
                        if flag/self.cnt >= self.min_support:
                            # 加入两者取并集后的项 和 支持度
                            sub_frequent_item.append(frequent_item[i] | frequent_item[j])
                            sub_support.append(list(set(support[i]).intersection(set(support[j]))))
        return sub_frequent_item, sub_support

    def findFrequentItem(self, frequent_item, support, frequent_set,support_set):
        # 先加入只包含1项 的 项集
        frequent_set.append(frequent_item)
        support_set.append(support)
        # 由k项构造k+1项 获得所有的频繁项集
        while len(frequent_item) >= 2 :
            frequent_item, support = self.getIntersection(frequent_item, support)
            frequent_set.append(frequent_item)
            support_set.append(support)

    # 对于每一个1项以上的频繁项集,都可能产生关联规则
    def generateRules(self, frequent_set, rules):
        for frequent_item in frequent_set:
            if len(frequent_item) > 1:
                self.getRules(frequent_item, frequent_item, frequent_set, rules)

    def removeItem(self, current_item, item):
        tempSet = []
        for elem in current_item:
            if elem != item:
                tempSet.append(elem)
        tempFrozenSet = frozenset(tempSet)
        return tempFrozenSet

    def getRules(self, frequent_item, current_item, frequent_set, rules):
        for item in current_item:
            # 去除item以计算置信度
            # 置信度 = 共同发生的概率 / 去除某个事件后发生的概率 例如 P(AB)/P(A)
            subset = self.removeItem(current_item, item)
            # print(subset)
            if frequent_set[subset]:
                confidence = frequent_set[frequent_item] / frequent_set[subset]
            else:
                confidence = 0
            # 如果大于最小置信度就存起来
            if confidence >= self.min_confidence:
                flag = False
                for rule in rules:# 判断是否已经存在相同规则
                    if (rule[0] == subset) and (rule[1] == frequent_item - subset):
                        flag = True

                if flag == False:# 没有则加入
                    rules.append((subset, frequent_item - subset, confidence))

                if len(subset) >= 2:# 递归挖掘更多规则
                    self.getRules(frequent_item, subset, frequent_set, rules)

    # 相当于主函数,Eclat算法完整过程
    def train(self, data, display=True):
        # 扫描数据集,反排获取垂直数据集
        frequent_item, support = self.invert(data)
        frequent_set = []
        support_set = []
        
        # 挖掘频繁项集
        self.findFrequentItem(frequent_item, support,frequent_set, support_set)

        # 将 出现事务TID 转换为 支持度
        data = {}
        for i in range(len(frequent_set)):
            for j in range(len(frequent_set[i])):
                data[frequent_set[i][j]] = len(support_set[i][j])

        rules = []
        # 挖掘关联规则
        self.generateRules(data, rules)
        # 输出结果
        workbook = openpyxl.load_workbook('Eclat.xlsx')
        sheet = workbook['Sheet1']
        for i in range(len(rules)):
            tmpcnt = 1
            for x in range(0,2):
                for j in rules[i][x]:
                    sheet.cell(i+1,tmpcnt).value = str(j)
                    tmpcnt = tmpcnt+1
            sheet.cell(i+1,tmpcnt).value = rules[i][2]
        workbook.save('Eclat.xlsx')
        return frequent_set, rules
        

if __name__ == "__main__":
    # 设置最小支持度,置信度,挖掘项数
    tmp = Eclat(0.01,0.9,3)
    # path引号内的改为处理后数据集的路径
    path = r"C:\Users\zbwer\Desktop\md\关联规则\source\超市关联规则数据集"
    os.chdir(path)
    # 加载数据集
    workbook = openpyxl.load_workbook('dataset.xlsx')
    sheet = workbook['Sheet3']
    rows= sheet.max_row+1
    cols = sheet.max_column+1
    NewDa = []
    for i in range(1,rows):
        tmpD = []
        for j in range(1,cols):
            celldate = sheet.cell(i,j).value
            if celldate:
                tmpD.append(celldate)
        NewDa.append(tmpD)
    tmp.cnt = float(len(NewDa))
    # 挖掘关联规则
    EndT1 = time.time()
    tmp.train(NewDa)
    # 记录程序运行时间
    EndT2 = time.time()
    print(EndT2-EndT1)
