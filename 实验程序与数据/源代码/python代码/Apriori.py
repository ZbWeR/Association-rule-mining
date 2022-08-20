import os
import openpyxl
import time
def load_data_set():
    # path引号内的改为处理后数据集的路径
    path = r"C:\Users\zbwer\Desktop\md\关联规则\source\超市关联规则数据集"
    os.chdir(path)

    workbook = openpyxl.load_workbook('dataset.xlsx')
    sheet = workbook['Sheet3']
    rows= sheet.max_row+1
    cols = sheet.max_column+1
    
    NewDa = []
    for i in range(1,rows):
        tmpD = []
        for j in range(1,cols):
            celldate = sheet.cell(i,j).value
            if celldate:
                tmpD.append(celldate)
        NewDa.append(tmpD)
    return NewDa

def create_C1(data_set):
    C1 = set()
    for t in data_set: # t相当于记录的编号
        for item in t:  # item相当于第t条记录中的所有内容,表现为数组
            item_set = frozenset([item]) # 将数组转化为集合
            C1.add(item_set) # 在C1候选集中加入 上述数据
    return C1

# 剪枝操作
def is_apriori(Ck_item, Lksub1):
    for item in Ck_item: # 取当前候选集中的 每一项
        sub_Ck = Ck_item - frozenset([item]) # 去除该项后得到一个k-1项的集合
        if sub_Ck not in Lksub1:# 判断这个集合是否属于k-1项频繁集
            return False
    return True


def create_Ck(Lksub1, k):
    # Lksub1 即k-1项的频繁集,由上一次的频繁集构造本次的候选集
    Ck = set()
    len_Lksub1 = len(Lksub1) # 频繁集个数
    list_Lksub1 = list(Lksub1) # 转为List类型
    for i in range(len_Lksub1):# 从0开始枚举所有的频繁集
        for j in range(1, len_Lksub1):# 从1开始枚举所有的频繁集
            # 取出当前频繁集的具体集合,表现为集合,此处转为数组储存
            l1 = list(list_Lksub1[i])
            l2 = list(list_Lksub1[j])
            # 排序,方便后序检验
            l1.sort()
            l2.sort()
            if l1[0:k-2] == l2[0:k-2]:# 如果除了最后一项前面都相同,说明可以构造
                Ck_item = list_Lksub1[i] | list_Lksub1[j] # 取集合的并集
                # pruning 剪枝
                if is_apriori(Ck_item, Lksub1):
                    Ck.add(Ck_item) # 构造候选集
    return Ck

# 使用候选集更新频繁集
def generate_Lk_by_Ck(data_set, Ck, min_support, support_data):
    Lk = set() # 存放k项频繁集
    item_count = {} 
    for t in data_set: # 枚举每一个事件
        for item in Ck: # 枚举每一个k项候选集
            if item.issubset(t): # 如果当前的候选集 是 事件的子集,计数器累加
                if item not in item_count:
                    item_count[item] = 1
                else:
                    item_count[item] += 1
    t_num = float(len(data_set)) # 事件的个数
    # 支持度 = 候选集出现次数/总事件个数
    for item in item_count:
        if (item_count[item] / t_num) >= min_support: # 大于最小支持度
            Lk.add(item) # 存入k项频繁集
            support_data[item] = item_count[item] / t_num # 同时记录该频繁集的支持度
    return Lk


def generate_L(data_set, k, min_support):
    support_data = {}
    C1 = create_C1(data_set) # 获得只有1项的候选集
    L1 = generate_Lk_by_Ck(data_set, C1, min_support, support_data) # 获得只有1项的频繁集
    Lksub1 = L1.copy()
    L = [] # 存放所有的频繁集,包括1项,2项...k项.
    L.append(Lksub1) # 放入1项频繁集
    for i in range(2, k+1):
        Ci = create_Ck(Lksub1, i) # 构造候选集
        Li = generate_Lk_by_Ck(data_set, Ci, min_support, support_data) # 构造频繁集
        Lksub1 = Li.copy() # 更新k-1项频繁集
        L.append(Lksub1) # 存入当前的频繁集
    # 需要注意的是 L 表现为一个三维的数组 L[i]表示 含有i项的频繁集
    # L[i][0]表示含有 i项的第一个频繁集 L[i][1]表示含有i项的第二个频繁集 一次类推
    # 每一个频繁集又可以看作一个一维数组
    return L, support_data # 返回所有频繁集及其对应的支持度


def generate_big_rules(L, support_data, min_conf):
    big_rule_list = []
    sub_set_list = []
    for i in range(0, len(L)): # 枚举频繁集所含的项数
        for freq_set in L[i]: # 枚举每一个含有i项的频繁集
            for sub_set in sub_set_list: # 枚举在此之前所有的频繁集
                if sub_set.issubset(freq_set):
            # 如果之前频繁集中有当前频繁集的子集 例如 [A]是[A,B]的子集
                    # 置信度 = 共同发生的概率 / 去除某个事件后发生的概率 例如 P(AB)/P(A)
                    conf = support_data[freq_set] / support_data[freq_set - sub_set]
                    big_rule = (freq_set - sub_set, sub_set, conf,support_data[freq_set])
                    if conf >= min_conf and big_rule not in big_rule_list:
                        big_rule_list.append(big_rule) # 满足最小置信度就加入规则中
            sub_set_list.append(freq_set)
    return big_rule_list

if __name__ == "__main__":
    ST = time.time()
    data_set = load_data_set() # 加载数据
    E1 = time.time()
    # 挖掘关联规则
    L, support_data = generate_L(data_set, k=3, min_support=0.01) # 寻找频繁集
    big_rules_list = generate_big_rules(L, support_data, min_conf=0.9) # 提取规则
    # 保存数据
    cnt = 0
    workbook = openpyxl.load_workbook('apriori.xlsx')
    sheet = workbook['Sheet1']
    for i in range(len(big_rules_list)):
        tmpcnt = 1
        for x in range(0,2):
            for j in big_rules_list[i][x]:
                sheet.cell(i+1,tmpcnt).value = str(j)
                tmpcnt = tmpcnt+1
        sheet.cell(i+1,tmpcnt).value = big_rules_list[i][2]
        sheet.cell(i+1,tmpcnt+1).value = big_rules_list[i][3]
    workbook.save('apriori.xlsx')
    #记录程序运行时间
    E2 = time.time()
    print(E2-E1)